import streamlit as st
import pdfplumber
import re
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Alignment
import matplotlib.pyplot as plt
import time

# -----------------------------
# Classification Function
# -----------------------------
def classify_student(records, percentage):
    subs_101_107 = [marks for subj, marks in records if subj.startswith("BED10") and int(subj[-3:]) <= 107]
    subs_108_112 = [marks for subj, marks in records if subj.startswith("BED10") and 108 <= int(subj[-3:]) <= 112]
    subs_201_205 = [marks for subj, marks in records if subj.startswith("BED20") and int(subj[-3:]) <= 205]
    subs_206_212 = [marks for subj, marks in records if subj.startswith("BED20") and 206 <= int(subj[-3:]) <= 212]

    # ---- First Year Rules ----
    if subs_101_107:
        if any(m == 0 for m in subs_101_107):
            return "Fail"
        count_below_50 = sum(1 for m in subs_101_107 if m < 50)
        if count_below_50 > 3:
            return "Fail"
        elif 1 <= count_below_50 <= 3:
            return "ATKT"
        if any(m == 0 for m in subs_108_112):
            return "Fail in Internal"

    # ---- Second Year Rules ----
    if subs_201_205:
        if any(m < 50 for m in subs_201_205):
            return "Fail"
    if subs_206_212:
        if any(m == 0 for m in subs_206_212):
            return "Fail in Internal"

    # ---- Percentage Classification ----
    if percentage >= 80:
        return "Distinction"
    elif percentage >= 65:
        return "First Class"
    elif percentage >= 55:
        return "Second Class"
    elif percentage >= 50:
        return "Pass"
    else:
        return "Fail"


# -----------------------------
# PDF Parser Function
# -----------------------------
def parse_pdf(file):
    students_data = {}
    current_student = None
    current_seat = None
    current_name = None

    with pdfplumber.open(file) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if not text:
                continue
            lines = text.split("\n")

            for line in lines:
                if line.startswith("PRN:"):
                    match = re.search(r"SEAT NO\.\:\s*(\d+)\s+NAME:\s*(.*?)\s+Mother", line)
                    if match:
                        current_seat = match.group(1).strip()
                        current_name = match.group(2).strip()
                        current_student = (current_seat, current_name)
                        students_data[current_student] = []

                match = re.match(r"(BED\s+\d{3}(-\d{2})?)\s+.*\s(\d{2,3})\s+\d+\s+\d+\s+\w+", line)
                if match and current_student:
                    subject = match.group(1).replace(" ", "")
                    marks = int(match.group(3))
                    students_data[current_student].append([subject, marks])

    return students_data


# -----------------------------
# Excel Builder Function
# -----------------------------
def build_excel(students_data):
    wb = Workbook()
    ws_all = wb.active
    ws_all.title = "All_Students"

    ws_all.append(["SEAT NO", "NAME", "Subject", "Marks", "Total", "Percentage", "Class"])

    row_num = 2
    summary_first_year = []
    summary_second_year = []
    all_classes = []
    fy_classes = []
    sy_classes = []

    for (seat_no, name), records in students_data.items():
        start_row = row_num
        total_marks = 0
        last_subject = None

        for subj, marks in records:
            total_marks += marks
            last_subject = subj
            ws_all.cell(row=row_num, column=1, value=seat_no if row_num == start_row else None)
            ws_all.cell(row=row_num, column=2, value=name if row_num == start_row else None)
            ws_all.cell(row=row_num, column=3, value=subj)
            ws_all.cell(row=row_num, column=4, value=marks)
            row_num += 1
        end_row = row_num - 1

        # Determine year & percentage
        if last_subject.endswith("112"):
            percentage = (total_marks / 1000) * 100
            classification = classify_student(records, percentage)
            summary_first_year.append([seat_no, name, total_marks, round(percentage, 2), classification])
            fy_classes.append(classification)
        elif last_subject.endswith("212"):
            percentage = (total_marks / 2000) * 100
            classification = classify_student(records, percentage)
            summary_second_year.append([seat_no, name, total_marks, round(percentage, 2), classification])
            sy_classes.append(classification)
        else:
            percentage = None
            classification = None

        all_classes.append(classification)

        # Write total, percentage, class
        ws_all.cell(row=end_row, column=5, value=total_marks)
        if percentage is not None:
            perc_cell = ws_all.cell(row=end_row, column=6, value=round(percentage, 2))
            perc_cell.number_format = "0.00"
        ws_all.cell(row=end_row, column=7, value=classification)

        # Merge seat no and name
        ws_all.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
        ws_all.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2)
        ws_all.cell(row=start_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws_all.cell(row=start_row, column=2).alignment = Alignment(horizontal="center", vertical="center")

    # Sort top 5 summaries
    summary_first_year = sorted(summary_first_year, key=lambda x: x[3], reverse=True)[:5]
    summary_second_year = sorted(summary_second_year, key=lambda x: x[3], reverse=True)[:5]

    # First Year sheet
    ws_first = wb.create_sheet("First Year")
    ws_first.append(["SEAT NO", "NAME", "Total", "Percentage", "Class"])
    for seat_no, name, total, perc, classification in summary_first_year:
        ws_first.append([seat_no, name, total, perc, classification])
        ws_first.cell(row=ws_first.max_row, column=4).number_format = "0.00"

    # Second Year sheet
    ws_second = wb.create_sheet("Second Year")
    ws_second.append(["SEAT NO", "NAME", "Total", "Percentage", "Class"])
    for seat_no, name, total, perc, classification in summary_second_year:
        ws_second.append([seat_no, name, total, perc, classification])
        ws_second.cell(row=ws_second.max_row, column=4).number_format = "0.00"

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output, summary_first_year, summary_second_year, all_classes, fy_classes, sy_classes


# -----------------------------
# Chart Plot Functions
# -----------------------------
def plot_class_distribution(counts, title, max_y=None):
    fig, ax = plt.subplots(figsize=(5, 3))   # ✅ Medium size
    bars = ax.bar(counts.index, counts.values, color="skyblue")

    for bar in bars:
        height = bar.get_height()
        ax.text(bar.get_x() + bar.get_width()/2, height,
                str(height), ha='center', va='bottom',
                fontsize=10, fontweight='bold')

    ax.set_title(title)
    ax.set_ylabel("Count")

    if max_y:
        ax.set_ylim(0, max_y)

    st.pyplot(fig)

    df_counts = counts.reset_index()
    df_counts.columns = ["Class", "Count"]
    df_counts.index = df_counts.index + 1
    df_counts.index.name = "S.No"   # ✅ Serial shown
    st.table(df_counts)


def plot_side_by_side(fy_counts, sy_counts):
    df_compare = pd.DataFrame({"First Year": fy_counts, "Second Year": sy_counts}).fillna(0)

    fig, ax = plt.subplots(figsize=(6, 4))   # ✅ Medium size
    df_compare.plot(kind="bar", ax=ax, width=0.7)

    for i, col in enumerate(df_compare.columns):
        for idx, val in enumerate(df_compare[col]):
            ax.text(idx + (i-0.5)*0.2, val, str(int(val)), ha='center', va='bottom', fontsize=9, fontweight="bold")

    ax.set_title("First Year vs Second Year — Class Distribution")
    ax.set_ylabel("Count")
    st.pyplot(fig)

    df_compare.index.name = "Class"
    df_compare = df_compare.reset_index()
    df_compare.index = df_compare.index + 1
    df_compare.index.name = "S.No"   # ✅ Serial shown
    st.table(df_compare)


# -----------------------------
# Streamlit UI
# -----------------------------
st.title("📊 Student Result Processor")
st.write("Upload University PDF Marksheet to extract and analyze results.")
st.write("Code Compile by Prof.Manoj Shinde (prof.manojshinde@gmail.com)")

uploaded_file = st.file_uploader("Upload PDF File", type=["pdf"])

if uploaded_file is not None:
    start_time = time.time()
    with st.spinner("⏳ Processing... please wait"):
        students_data = parse_pdf(uploaded_file)

    if not students_data:
        st.error("❌ No student data found in this PDF.")
    else:
        excel_file, summary_first_year, summary_second_year, all_classes, fy_classes, sy_classes = build_excel(students_data)

        runtime = time.time() - start_time
        st.info(f"⏱ Processing completed in {runtime:.2f} seconds.")

        df_first = pd.DataFrame(summary_first_year, columns=["SEAT NO", "NAME", "Total", "Percentage", "Class"])
        df_second = pd.DataFrame(summary_second_year, columns=["SEAT NO", "NAME", "Total", "Percentage", "Class"])
        df_first.index += 1
        df_second.index += 1

        st.subheader("🏆 First Year - Top 5 Students")
        st.table(df_first)

        st.subheader("🏆 Second Year - Top 5 Students")
        st.table(df_second)

        st.subheader("📊 Class Distribution (All Students)")
        class_counts = pd.Series(all_classes).value_counts().sort_index()
        plot_class_distribution(class_counts, "All Students", max_y=class_counts.max()+2)

        st.subheader("📊 Class Distribution — First Year")
        fy_counts = pd.Series(fy_classes).value_counts().sort_index()
        plot_class_distribution(fy_counts, "First Year", max_y=max(fy_counts.max(),1)+2)

        st.subheader("📊 Class Distribution — Second Year")
        sy_counts = pd.Series(sy_classes).value_counts().sort_index()
        plot_class_distribution(sy_counts, "Second Year", max_y=max(sy_counts.max(),1)+2)

        st.subheader("📊 First Year vs Second Year — Side-by-Side Comparison")
        plot_side_by_side(fy_counts, sy_counts)

        st.download_button(
            label="📥 Download Processed Excel",
            data=excel_file,
            file_name="all_final_result.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
